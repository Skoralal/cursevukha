import math
import numpy as np
from aiogram import Bot, Dispatcher , executor, types
import logging
import sqlite3
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import numbers
from openpyxl.styles.numbers import BUILTIN_FORMATS
TOKEN = "5736853255:AAH-03mQPiG7pCgYbrgdqq38F3UfUNDRNvs"
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
conn = sqlite3.connect("C:\prog_questionmark\cursevukha\logs.db")
cursor = conn.cursor()
wb = load_workbook("C:/prog_questionmark/test/data/samples/list.xlsx")
ws = wb.active




@dp.message_handler(commands=["start", "help"])
async def start(message : types.Message):
    await bot.send_message(message.from_user.id, "В боте представлен функционал по всем задачам курсовой. Cписок команд: /first, /second_zero, /second_first, /second_second, /third, /fourth, /fifth")
    await bot.send_message(message.from_user.id, "Чтобы ввести свои данные нужно указать их в одном сообщение вместе с командой. В каждом задании представлены разные группы данных например в первом задании нужно ввести 50(может меняться) чисел и указать количество групп.")
    await bot.send_message(message.from_user.id, "Пример команды для каждого задания. можно использовать точки вместо запятых")
    await bot.send_message(message.from_user.id, "Пример команды для каждого задания. можно использовать точки вместо запятых")
    await bot.send_message(message.from_user.id, "/first 19000 41000 90000 35000 59000 43000 70000 41000 49000 59000 59000 44000 12000 25000 60000 71000 43000 10000 63000 59000 78000 44000 17000 41000 59000 55000 20000 20000 65000 59000|9")
    await bot.send_message(message.from_user.id, "/second_zero 2001 2002 2003 2004 2005 2006 2007 2008 2009|890 870 980 960 1150 1150 1260 1210 1240|2023|2024")
    await bot.send_message(message.from_user.id, "/second_first 2003 2004 2005 2006 2007 2008 2009 2010|640 599 647 601|655 612 670 630 685")
    await bot.send_message(message.from_user.id, "/second_second 2001 2002 2003 2004 2005|129 110,4 100 123,4 130|50 60 80,6 100 140,1")
    await bot.send_message(message.from_user.id, "/third 2008 2009 2010|205 174 163 326 326 313 136 131 128 336 339 348|168 151 163 330 326 319 138 132 128 340 341 351|154 146 156 322 326 315 134 129 128 338 345 349")
    await bot.send_message(message.from_user.id, "/fourth 130 129 128 127,9 127,8 127,6 127,5 127,4 127,3 127,1|12 10,8 8,2 7,56 9,17 10 8,6 9,21 11 7,98")
    await bot.send_message(message.from_user.id, "/fifth 8070 105 43 32,78 44,6")
    await bot.send_message(message.from_user.id, "Если ты отправил сообщение, а бот молчит, есть 2 причины. 1) Бот в оффлайне. Проверить это можно, написав /start или /help, тут то ошибиться в написании сложно. 2) Ошибся в написании. Не там поставил | не так указал порядок, не указал все аргументы и так далее. Переделывай. 3) Вообще есть ещё 1. Напишите /complaint и вставьте полную команду. Я посмотрю")
    print(message.text)
    try:
        cursor.execute("INSERT INTO `users` (`user_id`, `user_first_name`, `user_last_name`) VALUES (?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name))
        conn.commit()
    except:
        print("aboba")    
    cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
    conn.commit()


@dp.message_handler(commands=["first"])
async def first(message : types.message):
    try:
        print(message.text)
        i1_input = message.text
        i1_cmd = []
        i1_cmd = i1_input.split("|")
        #input_user = input("Вставь все значения из 1 задания через пробел: ")
        #row = input_user.split(" ")
        row = []
        row_prep = i1_cmd[0]
        row = row_prep.split(" ")
        row.remove("/first")
        i = 0
        int_row = []
        for value in row:
            i += 1
            int_row.append(int(value))
        await bot.send_message(message.from_user.id, f"Первоначальный ряд {int_row}")
        await bot.send_message(message.from_user.id, f"Число наблюдений N={i}")
        int_row.sort()
        await bot.send_message(message.from_user.id, f"Ранжированный ряд {int_row}")
        #print(int_row)
        for value_2 in int_row:
            freq = int_row.count(value_2)
        #-----------------------------------
        disc_row = []
        answer_1_3 = []
        disc_freq_list = []
        disc_freq_list_2 = []
        act_disc_row = []
        for value in int_row:
            if value not in act_disc_row:
                act_disc_row.append(value)
        act_count = []
        for inde, value in enumerate(act_disc_row):
            act_count.append(int_row.count(value))
        for value_2 in int_row:
            if value_2 not in disc_row:
                disc_freq = int_row.count(value_2)
                disc_row.append(f"{value_2} - {disc_freq}")
        for value_3 in disc_row:
            if value_3 not in answer_1_3:
                answer_1_3.append(value_3)
        #print(answer_1_3)
        await bot.send_message(message.from_user.id, f"Дискретный ряд {answer_1_3}")
        disc_freq_list_prep = str(answer_1_3)
        disc_freq_list_prep_0 = disc_freq_list_prep.replace(",", ",")
        disc_freq_list_prep_1 = disc_freq_list_prep_0.replace("'[", "")
        disc_freq_list_prep_2 = disc_freq_list_prep_1.replace("]'", "")
        disc_freq_list_prep_3 = disc_freq_list_prep_2.replace("' '", "")
        disc_freq_list_prep_4 = disc_freq_list_prep_3.replace("['", "")
        disc_freq_list_prep_5 = disc_freq_list_prep_4.replace("']", "")
        disc_freq_list_prep_6 = disc_freq_list_prep_5.replace("', '", " - ")
        #await bot.send_message(message.from_user.id, f" disc_freq_list_prep {disc_freq_list_prep_6}")
        disc_freq_list = disc_freq_list_prep_6.split(" - ")
        #await bot.send_message(message.from_user.id, f" {disc_freq_list}")
        for ind, value in enumerate(disc_freq_list):
            if ind % 2 != 0:
                disc_freq_list_2.append(value)
        #await bot.send_message(message.from_user.id, f" {disc_freq_list_2}")
        disc_sum = []
        disc_sum_answer = ""
        for value in disc_freq_list_2:
            disc_sum.append(value)
            disc_sum.append("+")
        disc_sum = disc_sum[0:-1]
        for value in disc_sum:
            disc_sum_answer += value
        await bot.send_message(message.from_user.id, f"Проверка общей суммы {disc_sum_answer} = {i}")
        #-----------------------------------
        #group_q = int(input("Скока групп в варианте? "))
        group_q = int(i1_cmd[1])
        wb = load_workbook(f"C:/prog_questionmark/test/data/samples/list1_{group_q}.xlsx")
        ws = wb.active
        max_1 = max(int_row)
        min_1 = min(int_row)
        interval = round(math.ceil((max_1 - min_1)/group_q) / 50) * 50
        if min_1 + group_q * interval < max_1:
            interval += 50
        await bot.send_message(message.from_user.id, f"Х макс {max_1}")
        await bot.send_message(message.from_user.id, f"Х мин {min_1}")
        await bot.send_message(message.from_user.id, f"Интервал - {interval}")
        #------------------------------------
        await bot.send_message(message.from_user.id, "Таблица. Первый столбец - значения интервалов. Второй - частота. Третий - середина интервала. Четвёртый - середина интервала, умноженная на частоту. Пятый - порядковый номер.")
        lower_t_list = []
        upper_t_list = []
        freq_2_list = []
        table_1 = []
        avg_1_row = []
        xi_fi = []
        table_1_rows = 0
        temp_counter = 0
        while table_1_rows < group_q:
            lower_t = min_1 + interval * table_1_rows
            lower_t_list.append(lower_t)
            upper_t = min_1 + interval * table_1_rows + interval
            upper_t_list.append(upper_t)
            freq_2 = 0
            temp_counter += 1
            for value_3 in int_row:
                if value_3 > lower_t and value_3 <= upper_t:#было больше равно и меньше
                    freq_2 += 1
                elif value_3 == lower_t and 1 == temp_counter :#было upper вместо lower и 1 вместо group_q
                    freq_2 += 1
            freq_2_list.append(freq_2)
            try:
                avg_1 = int((lower_t + upper_t) / 2)
            except:
                avg_1 = (lower_t + upper_t) / 2
            avg_1_row.append(avg_1)
            xi_fi.append(avg_1 * freq_2)
            await bot.send_message(message.from_user.id, (f"{lower_t} - {upper_t} -|- {freq_2} -|- {avg_1} -|- {avg_1 * freq_2} -|- {table_1_rows+1}"))
            table_1_rows += 1
        #---------------------------------------
        avg_1_1_row = []
        for i_3, value in enumerate(avg_1_row):
            avg_1_1_row.append(value * freq_2_list[i_3])
        avg_1_1 = sum(int_row) / i
        sum_int_row = sum(int_row)
        await bot.send_message(message.from_user.id, f"Это числитель {sum_int_row}")
        await bot.send_message(message.from_user.id, f"Это знаменатель {i}")
        await bot.send_message(message.from_user.id, f"Средневзвешенная - {avg_1_1}")
        mid_value = []
        for v1, v2 in zip(avg_1_row, freq_2_list):
            mid_value.append(v1 * v2)
        mid_avg = sum(avg_1_1_row) / i
        await bot.send_message(message.from_user.id, f"Числитель {sum(avg_1_1_row)}")
        await bot.send_message(message.from_user.id, f"Знаменатель {i}")
        await bot.send_message(message.from_user.id, f"Средневзвешенная для интервальноо ряда - {mid_avg}")
        #---------------------------------------
        index_max_1 = freq_2_list.index(max(freq_2_list))
        xm0 = lower_t_list[index_max_1]
        h = interval
        mm0 = freq_2_list[index_max_1]
        mm0_minus_1 = freq_2_list[index_max_1 - 1]
        mm0_plus_1 = freq_2_list[index_max_1 + 1]
        m0 = xm0 + h*((mm0 - mm0_minus_1) / ((mm0 - mm0_minus_1) + (mm0 - mm0_plus_1)))
        await bot.send_message(message.from_user.id, f"Хм0 = {xm0}")
        await bot.send_message(message.from_user.id, f"h = {h}")
        await bot.send_message(message.from_user.id, f"mm0 = {mm0}")
        await bot.send_message(message.from_user.id, f"mm0-1 = {mm0_minus_1}")
        await bot.send_message(message.from_user.id, f"mm0+1 = {mm0_plus_1}")
        await bot.send_message(message.from_user.id, f"Мода = {round(m0, 2)}")
        moda_index = 0
        moda_number = 0
        for inde, value in enumerate(act_count):
            if value > moda_number:
                moda_number = value
                moda_index = inde
        #for inde, value in enumerate(act_count):
            #if value > moda_index:
                #moda_index = act_disc_row[inde]
        moda = act_disc_row[moda_index]
        act_moda = 0
        for modval in int_row:
            if int_row.count(modval)>act_moda:
                act_moda = modval
        await bot.send_message(message.from_user.id, f"Мода дискретного ряда = {moda}")
        #--------------------------------------
        mid_1_1_index = int(len(int_row) / 2)
        mid_1_2_index = int(len(int_row) / 2 - 1)
        mid_1_1 = int_row[mid_1_1_index]
        mid_1_2 = int_row[mid_1_2_index]
        mid_1 = (mid_1_1 + mid_1_2) / 2
        index_moda = 0
        for ima in range(len(lower_t_list)-1):
            if lower_t_list[ima]<=moda:
                index_moda = ima
        # index_moda += 1
        index_moda = freq_2_list.index(max(freq_2_list))
        await bot.send_message(message.from_user.id, f"xm = {mid_1_1}")
        await bot.send_message(message.from_user.id, f"xm + 1 = {mid_1_2}")
        await bot.send_message(message.from_user.id, "2, просто 2")
        await bot.send_message(message.from_user.id, f"Медиана для дискретного ряда = {mid_1}")
        #await bot.send_message(message.from_user.id, f"Медиана = {max(int_row)}")
        index_mid = math.floor(group_q / 2)
        vme_minus_1 = 0
        temp_4 = 0
        for value in freq_2_list:
            if temp_4 < index_mid:
                vme_minus_1 += value
                temp_4 += 1
        print(index_moda)
        print(222)
        for i_mat, value_mat in enumerate(freq_2_list):
            if sum(freq_2_list[:i_mat]) > i/2:
                mat_ebal_index = i_mat
                break
        print(mat_ebal_index)
        mat_ebal_index -= 1
        int_row_mid = round(lower_t_list[index_mid] + interval * ((i / 2 - vme_minus_1) / freq_2_list[index_mid]), 2)
        await bot.send_message(message.from_user.id, f"Xmin = {lower_t_list[mat_ebal_index]}")
        await bot.send_message(message.from_user.id, f"Первое значение в числителе{i / 2}")

        await bot.send_message(message.from_user.id, f"Vme - 1 = {sum(freq_2_list[0:mat_ebal_index])}")
        await bot.send_message(message.from_user.id, f"Fme = {freq_2_list[mat_ebal_index]}")
        await bot.send_message(message.from_user.id, f"Медиана для интервального ряда {lower_t_list[mat_ebal_index]+interval*((((i/2)-(sum(freq_2_list[0:mat_ebal_index])))/freq_2_list[mat_ebal_index]))}")
        #----------------------------------------
        razmakh = max_1 - min_1
        await bot.send_message(message.from_user.id, f"Rmax {max_1}")
        await bot.send_message(message.from_user.id, f"Rmin {min_1}")
        await bot.send_message(message.from_user.id, f"Размах вариации = {razmakh}")
        #----------------------------------------
        #deviation = []
        avg_2_row = []
        x = []
        for avg_2 in avg_1_row:
            avg_2_row.append(avg_2)
        itt_1 = 0
        itt_row_1 = []
        for avg_2 in avg_2_row:
            itt_row_1.append(avg_2 * freq_2_list[itt_1])
            itt_1 += 1
        itt_val_sum = 0
        for itt_val in itt_row_1:
            itt_val_sum += itt_val
        x_s_palkoi = round(itt_val_sum / i, 2)
        itt_2 = 0
        temp = 0
        temp_1 = 0
        itt_2_row = []
        square_row = []
        for value in avg_2_row:
            temp_1 = round(value - x_s_palkoi)
            if temp_1 < 0:
                square_row.append(temp_1 * -1)
            else:
                square_row.append(temp_1)
        for value in freq_2_list:
            temp = round(value * (avg_2_row[itt_2] - x_s_palkoi), 2)
            if temp < 0:
                itt_2_row.append(temp * -1)
            else:
                itt_2_row.append(temp)
            itt_2 += 1
        itt_2_row_sum = 0
        for value in itt_2_row:
            itt_2_row_sum += value
        square_row_fr = []
        itt_2_row_2 = []
        for value in square_row:
            square_row_fr.append(value * value)
        temp_3 = 0
        for value in square_row_fr:
            itt_2_row_2.append(value * freq_2_list[temp_3])
            temp_3 += 1
        itt_2_row_2_sum = 0
        for value in itt_2_row_2:
            itt_2_row_2_sum += value
        avg_line = round(itt_2_row_sum / i, 2)
        #await bot.send_message(message.from_user.id, f"Среднее линейное отклонение - {avg_line}")
        disp = round(itt_2_row_2_sum / i, 2)
        avg_quad = round(disp ** 0.5, 2)
        #await bot.send_message(message.from_user.id, f"Среднее квадратичное отклонение - {avg_quad}")
        #await bot.send_message(message.from_user.id, f"Дисперсия - {disp}")
        #-------------------------------
        osc = round(razmakh * 100 / x_s_palkoi, 2)
        #await bot.send_message(message.from_user.id, f"Коэффициент осциляции - {osc}%")
        #-------------------------------------
        v_ro = round(avg_quad * 100 / x_s_palkoi, 2)
        #await bot.send_message(message.from_user.id, f"Коэффициент вариации - {v_ro}%")
        #-----------------------------------------
        #await bot.send_message(message.from_user.id, f"Линейный коэффициент вариации = {round(avg_line * 100 / x_s_palkoi, 2)}%")
        await bot.send_message(message.from_user.id, f"Вспомогательная таблица")
        freq_add = 0
        freq_add_row = []
        prelast = []
        last = []
        for inde, value in enumerate(freq_2_list):
            freq_add += value
            freq_add_row.append(freq_add)
            try:
                prelast.append(int(abs((avg_1_row[inde] - mid_avg) * value)))
            except:
                prelast.append(abs((avg_1_row[inde] - mid_avg) * value))
            try:
                last.append(int(abs((avg_1_row[inde] - mid_avg) ** 2 * value)))
            except:
                last.append(abs((avg_1_row[inde] - mid_avg) ** 2 * value))
            await bot.send_message(message.from_user.id, (f"{inde+1} -|- {freq_add} -|- {avg_1_row[inde] * freq_2_list[inde]} -|- {abs((avg_1_row[inde] - mid_avg) * value)} -|- {abs((avg_1_row[inde] - mid_avg) ** 2 * value)}"))
        minus_2_col = []
        for inde, value in enumerate(avg_1_row):
            minus_2_col.append(value * freq_2_list[inde])
        minus_2_col_avg = sum(minus_2_col) / i
        await bot.send_message(message.from_user.id, f"Числитель {itt_2_row_sum}")
        await bot.send_message(message.from_user.id, f"Знаменатель {i}")
        await bot.send_message(message.from_user.id, f"Среднее линейное отклонение - {avg_line}")
        await bot.send_message(message.from_user.id, f"Под корнем {disp}")
        await bot.send_message(message.from_user.id, f"Среднее квадратичное отклонение - {avg_quad}")
        minus_1_col = []
        for inde, value in enumerate(avg_1_row):
            minus_1_col.append(value )
        await bot.send_message(message.from_user.id, f"Числитель {itt_2_row_2_sum}")
        await bot.send_message(message.from_user.id, f"Знаменатель {i}")
        await bot.send_message(message.from_user.id, f"Дисперсия - {disp}")
        await bot.send_message(message.from_user.id, f"Числитель {razmakh}")
        await bot.send_message(message.from_user.id, f"Знаменатель {mid_avg}")
        await bot.send_message(message.from_user.id, f"Коэффициент осциляции - {osc}%")
        await bot.send_message(message.from_user.id, f"Числитель {avg_quad}")
        await bot.send_message(message.from_user.id, f"Знаменатель {mid_avg}")
        await bot.send_message(message.from_user.id, f"Коэффициент вариации - {v_ro}%")
        await bot.send_message(message.from_user.id, f"Числитель {avg_line}")
        await bot.send_message(message.from_user.id, f"Знаменатель {mid_avg}")
        await bot.send_message(message.from_user.id, f"Линейный коэффициент вариации = {round(avg_line * 100 / x_s_palkoi, 2)}%")
        #{abs((avg_1_row[inde] - mid_avg) * value)} -|- {abs((avg_1_row[inde] - mid_avg) ** 2 * value)}
        for q, value in enumerate(range(group_q)):
            ws[f"A{2+q}"].value = f"{lower_t_list[q]} - {upper_t_list[q]}"
        for q, value in enumerate(range(group_q)):
            ws[f"B{2+q}"].value = freq_2_list[q]
        for q, value in enumerate(range(group_q)):
            ws[f"C{2+q}"].value = avg_1_row[q]
        for q, value in enumerate(range(group_q)):
            ws[f"D{2+q}"].value = xi_fi[q]
        for q, value in enumerate(range(group_q)):
            ws[f"D{2+q}"].value = xi_fi[q]
        for q, value in enumerate(range(group_q)):
            ws[f"E{2+q}"].value = prelast[q]
        for q, value in enumerate(range(group_q)):
            ws[f"F{2+q}"].value = last[q]
        #for q, value in enumerate(range(group_q)):
            #ws[f"B{2+group_q+2+q}"].value = 0
            #ws[f"B{2+group_q+2+q}"].number_format = numbers.FORMAT_DATE_DATETIME
        wb.save("C:/prog_questionmark/test/data/first_first/list2281488.xlsx")
        for q, value in enumerate(range(group_q)):
            if q == group_q-1:
                ws[f"B{2+group_q+2+q}"].value = freq_add_row[q] * 365.25 + freq_add_row[q]
                ws[f"G{2+group_q+2+q}"].value = f"=B{2+group_q+2+q}"
                ws[f"G{2+group_q+2+q}"].number_format = 'YY'
                ws[f"B{2+group_q+2+q}"].number_format = 'YY'
                ws[f"B{2+group_q+2+q+1}"].value = 0
                ws[f"G{2+group_q+2+q+1}"].value = 0
                ws[f"G{2+group_q+2+q}"].number_format = 'YY'
                ws[f"G{2+group_q+2+q+1}"].number_format = 'YY'
                ws[f"D{2+group_q+2+q}"].value = avg_1_row[q]
                ws[f"I{2+group_q+2+q}"].value = f"=D{2+group_q+2+q}"
            else:
                ws[f"B{2+group_q+2+q}"].value = freq_add_row[q] * 365.25 + freq_add_row[q]
                ws[f"B{2+group_q+2+q}"].number_format = 'YY'
                ws[f"D{2+group_q+2+q}"].value = avg_1_row[q]
                ws[f"I{2+group_q+2+q}"].value = f"=D{2+group_q+2+q}"
                ws[f"G{2+group_q+2+q}"].value = f"=B{2+group_q+2+q}"
                ws[f"G{2+group_q+2+q}"].number_format = 'YY'
        for q, value in enumerate(range(group_q)):
            ws[f"B{2+group_q+3+q+group_q}"].value = freq_add_row[q] * 365.25 + freq_add_row[q]
            ws[f"B{2+group_q+3+q+group_q}"].number_format = 'YY'
            ws[f"G{2+group_q+3+q+group_q}"].value = f"=B{2+group_q+3+q+group_q}"
            ws[f"G{2+group_q+3+q+group_q}"].number_format = 'YY'
            ws[f"G{2+group_q+3+q+group_q+2+group_q}"].value = upper_t_list[q]
            ws[f"H{2+group_q+3+q+group_q+2+group_q}"].value = freq_2_list[q]
            ws[f"I{2+group_q+3+q+group_q+2+group_q}"].value = freq_2_list[q]
            ws[f"D{2+group_q+3+q+group_q}"].value = avg_1_row[q]
            ws[f"I{2+group_q+3+q+group_q}"].value = f"=D{2+group_q+3+q+group_q}"
        chart1 = BarChart()
        chart1.add_data(f"Лист1!H{2+group_q+3+0+group_q+group_q+2}:H{2+group_q+3+group_q+group_q+group_q+1}")
        abob = Reference(ws, min_col=7, min_row=2+group_q+3+0+group_q+group_q+2, max_row=2+group_q+3+group_q+group_q+group_q+1)
        chart1.set_categories(abob)
        chart2 = LineChart()
        chart2.add_data(f"Лист1!I{2+group_q+3+0+group_q+group_q+2}:I{2+group_q+3+group_q+group_q+group_q+1}")
        chart1 += chart2
        ws.add_chart(chart1)
        for q, value in enumerate(range(group_q)):
            ws[f"C{2+group_q+2}"].value = lower_t_list[0]
            ws[f"H{2+group_q+2}"].value = f"=C{2+group_q+2}"
        for q, value in enumerate(range(group_q)):
            if q != group_q-1:
                ws[f"C{2+group_q+3+q}"].value = upper_t_list[q]
                ws[f"H{2+group_q+3+q}"].value = f"=C{2+group_q+3+q}"
            else:
                ws[f"C{2+group_q+3+q}"].value = lower_t_list[0]
                ws[f"H{2+group_q+3+q}"].value = f"=C{2+group_q+3+q}"
        for q, value in enumerate(range(group_q)):
            ws[f"C{2+group_q+2+group_q+q+1}"].value = upper_t_list[q]
            ws[f"H{2+group_q+2+group_q+q+1}"].value = f"=C{2+group_q+2+group_q+q+1}"
        data_1 = ws["I2:R6"]
        data_counter = 0
        for row_1 in data_1:
            for col in row_1:
                try:
                    col.value = row[data_counter]
                    data_counter +=1
                except:
                    break

        
        #for q, value in enumerate(range(group_q)):
            #ws[f"D{2+group_q+3+q}"].value = avg_1_row[q]
        #for q, value in enumerate(range(group_q)):
            #ws[f"E{2+group_q+3+q}"].value = freq_2_list[q]

        wb.save(f"C:/prog_questionmark/test/data/first_first/first{message.chat.id}.xlsx")
        await bot.send_document(message.chat.id, open(f"C:/prog_questionmark/test/data/first_first/first{message.chat.id}.xlsx", 'rb'))
        await bot.send_message(message.from_user.id, "Графики нужно дооформить")
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["second_zero"])
async def second_zero(message : types.message):
    try:
        #enumerate
        wb = load_workbook("C:/prog_questionmark/test/data/samples/list.xlsx")
        ws = wb.active  
        i2_0_input = message.text
        i2_0_cmd = []
        i2_0_cmd = i2_0_input.split("|")
        years_input = i2_0_cmd[0]
        #years_input = input("Вставь годы через пробел ")
        workers_input = i2_0_cmd[1]
        #workers_input = input("А сюда количество работников через пробел ")
        years_str = years_input.split(" ")
        years_str.remove("/second_zero")
        workers_input = workers_input.replace(",", ".")
        workers_str = workers_input.split(" ")
        #TEMP
        #years_str = ['2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017']
        #workers_str = ['590', '510', '475', '512', '568', '468', '643', '610', '662', '620']
        #years_str = ['2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009']
        #workers_str = ['890', '870', '980', '960', '1150', '1150', '1260', '1210', '1240']
        #TEMP
        count = len(years_str)
        years = []
        workers = []
        for value in years_str:
            years.append(int(value))
        for value in workers_str:
            workers.append(float(value)) 
        abs_gr_chain = []
        abs_gr_base = []
        for i, value in enumerate(years):
            if i > 0:
                abs_gr_chain.append(workers[i] - workers[i-1])
            else:
                abs_gr_chain.append(0)
        for i, value in enumerate(years):
            if i > 0:
                abs_gr_base.append(workers[i] - workers[0])
            else:
                abs_gr_base.append(0)
        temp_r_chain = []
        temp_r_base = []
        for i, value in enumerate(years):
            if i > 0:
                temp_r_chain.append(round(workers[i] * 100 / workers[i-1], 2))
            else:
                temp_r_chain.append(100)
        for i, value in enumerate(years):
            if i > 0:
                temp_r_base.append(round(workers[i] * 100 / workers[0], 2))
            else:
                temp_r_base.append(100)
        temp_pr_chain = []
        temp_pr_base = []
        for i, value in enumerate(temp_r_chain):
            if i > 0:
                temp_pr_chain.append(round(temp_r_chain[i] - 100, 2))
            else:
                temp_pr_chain.append(0)
        for i, value in enumerate(temp_r_base):
            if i > 0:
                temp_pr_base.append(round(temp_r_base[i] - 100, 2))
            else:
                temp_pr_base.append(0)
        for i, year in enumerate(years):
            await bot.send_message(message.from_user.id, f"{year} ||| {workers[i]} ||| {abs_gr_chain[i]} ||| {abs_gr_base[i]} ||| {temp_r_chain[i]} ||| {temp_r_base[i]} ||| {temp_pr_chain[i]} ||| {temp_pr_base[i]}")
        workers_sum = 0
        for value in workers:
            workers_sum += value
        await bot.send_message(message.from_user.id, f"Числитель {workers_sum}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count}")
        await bot.send_message(message.from_user.id, f"Средний уровень ряда {round(workers_sum / count ,2)} чел.")
        #------------------------------------------------
        last_change = abs_gr_base[-1]
        await bot.send_message(message.from_user.id, f"Числитель {last_change}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count - 1}")
        await bot.send_message(message.from_user.id, f"Средний абсолютный прирост {round(last_change / (count - 1), 2)}")
        #-------------------------------------------------
        one_stepen = 1 / (count - 1)
        first_workers = workers[0]
        last_workers = workers[-1]
        await bot.send_message(message.from_user.id, f"Степень корня {count - 1}")
        await bot.send_message(message.from_user.id, f"Числитель {last_workers}")
        await bot.send_message(message.from_user.id, f"Знаменатель {first_workers}")
        await bot.send_message(message.from_user.id, f"Средный темп роста {round(((last_workers / first_workers) ** one_stepen) * 100, 2)}%")
        #--------------------------------------------------
        await bot.send_message(message.from_user.id, f"То же самое, но в конце -100")
        await bot.send_message(message.from_user.id, f"Средний темп прироста {round(((last_workers / first_workers) ** one_stepen) * 100 - 100, 2)}%")
        #--------------------------------------------------
        three_year_avg = 0
        three_year_avg_row = []
        three_year_row = []
        for i, value in enumerate(workers):
            if i < 1 or i == count - 1:
                three_year_avg_row.append(None)
                three_year_row.append(0)
            else:
                three_year_avg_row.append(round((workers[i] + workers[i-1] + workers[i+1]) / 3, 2))
                three_year_row.append(workers[i] + workers[i-1] + workers[i+1])
        #------------------------------------------
        #------------------------------------------
        x_s_palkoi = sum(workers) / count
        if count % 2 == 0:
            t = 2
        else:
            t = 1
        t_row = []
        t_base = 0
        if t == 2:
            t_base = count - 1
        else:
            t_base = math.floor(count / 2)
        #print(t_base)
        for value in range(count):
            t_row.append(t_base * -1 + t * value)
        t_row_sqrt = []
        for value in t_row:
            t_row_sqrt.append(value * value)
        flat_row = []
        sum_y_t = []
        for i, value in enumerate(workers):
            sum_y_t.append(workers[i] * t_row[i])
        sum_t_sqrt = sum(t_row_sqrt)
        b = round(sum(sum_y_t) / sum_t_sqrt, 2)
        #----------------------------------------------
        y_s_palkoi = []
        for i, value in enumerate(range(count)):
            y_s_palkoi.append(round(x_s_palkoi + b * t_row[i], 0))
        #print(y_s_palkoi)
        y_minus_1 = []
        for i, value in enumerate(years):
            y_minus_1.append(workers[i] - y_s_palkoi[i])
        #print(y_minus_1)
        y_minus_1_2 = []
        for value in y_minus_1:
            y_minus_1_2.append(value * value)
        #print(y_minus_1_2)
        disp = sum(y_minus_1_2) / count
        await bot.send_message(message.from_user.id, f"Суммы {three_year_row}")
        await bot.send_message(message.from_user.id, f"3-летняя средняя {three_year_avg_row}")
        for i,value in enumerate(years):
            await bot.send_message(message.from_user.id, f"{value} ||| {t_row[i]} ||| {sum_y_t[i]} ||| {t_row_sqrt[i]} ||| {y_s_palkoi[i]} ||| {y_minus_1[i]} ||| {y_minus_1_2[i]}")
        await bot.send_message(message.from_user.id, f"Числитель {workers_sum}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count}")
        await bot.send_message(message.from_user.id, f"a = {workers_sum / count}")
        await bot.send_message(message.from_user.id, f"Числитель {sum(sum_y_t)}")
        await bot.send_message(message.from_user.id, f"Знаменатель {sum_t_sqrt}")
        await bot.send_message(message.from_user.id, f"b = {sum(sum_y_t) / sum_t_sqrt}")
        await bot.send_message(message.from_user.id, f"Числитель {sum(y_minus_1_2)}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count}")
        await bot.send_message(message.from_user.id, f"Дисперсия {disp}")
        avg_quad_disp = round(disp ** 0.5, 2)
        await bot.send_message(message.from_user.id, f"Под корнем {disp}")
        await bot.send_message(message.from_user.id, f"Среднее квадратичное отклонения {avg_quad_disp}")
        await bot.send_message(message.from_user.id, f"Числитель {avg_quad_disp}")
        await bot.send_message(message.from_user.id, f"Знаменатель {x_s_palkoi}")
        await bot.send_message(message.from_user.id, f"Коэффициент вариации {round(avg_quad_disp * 100 / x_s_palkoi, 2)}%")
        #------------------------------------------
        future_year_1 = int(i2_0_cmd[2])
        future_year_2 = int(i2_0_cmd[3])
        #future_year_1 = int(input("Первый прогнозируемый год "))
        #future_year_2 = int(input("Второй прогнозируемый год "))
        #TEMP
        #future_year_1 = 2019
        #future_year_2 = 2020
        #TEMP
        gap_1 = future_year_1 - years[-1]
        gap_2 = future_year_2 - years[-1]
        future_ans_1 = (t_row[-1] + gap_1 * t) * b + x_s_palkoi
        future_ans_2 = (t_row[-1] + gap_2 * t) * b + x_s_palkoi
        await bot.send_message(message.from_user.id, f"{x_s_palkoi} + {b} * {t_row[-1] + gap_1 * t} = {future_ans_1}")
        await bot.send_message(message.from_user.id, f"В {future_year_1} - {future_ans_1}")
        await bot.send_message(message.from_user.id, f"{x_s_palkoi} + {b} * {t_row[-1] + gap_2 * t} = {future_ans_2}")
        await bot.send_message(message.from_user.id, f"В {future_year_2} - {future_ans_2}")
        #------------------------------------------
        for i, value in enumerate(years):
            ws[f"A{2+i}"].value = years[i]
        for i, value in enumerate(years):
            ws[f"B{2+i}"].value = workers[i]
        for i, value in enumerate(years):
            ws[f"C{2+i}"].value = abs_gr_chain[i]
        for i, value in enumerate(years):
            ws[f"D{2+i}"].value = f"{abs_gr_base[i]}"
        for i, value in enumerate(years):
            ws[f"E{2+i}"].value = f"{temp_r_chain[i]}%"
        for i, value in enumerate(years):
            ws[f"F{2+i}"].value = f"{temp_r_base[i]}%"
        for i, value in enumerate(years):
            ws[f"G{2+i}"].value = f"{temp_pr_chain[i]}%"
        for i, value in enumerate(years):
            ws[f"H{2+i}"].value = f"{temp_pr_base[i]}%"
        for i, value in enumerate(years):
            ws[f"I{2+i}"].value = y_s_palkoi[i]
        for i, value in enumerate(years):
            ws[f"J{2+i}"].value = three_year_avg_row[i]
            ws[f"O{2+i}"].value = value
            ws[f"P{2+i}"].value = t_row[i]
            ws[f"Q{2+i}"].value = sum_y_t[i]
            ws[f"R{2+i}"].value = t_row_sqrt[i]
            ws[f"S{2+i}"].value = y_s_palkoi[i]
            ws[f"T{2+i}"].value = y_minus_1[i]
            ws[f"U{2+i}"].value = y_minus_1_2[i]
        #f"{value} ||| {t_row[i]} ||| {sum_y_t[i]} ||| {t_row_sqrt[i]} ||| {y_s_palkoi[i]} ||| {y_minus_1[i]} ||| {y_minus_1_2[i]}")
        mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        table = openpyxl.worksheet.table.Table(ref=f'A1:J{2+i}', displayName='table2', tableStyleInfo=mediumStyle)
        ws.add_table(table)
        chart = LineChart()
        chart.add_data(f"Лист1!B2:B{2+i}")
        chart.add_data(f"Лист1!I2:I{2+i}")
        chart.add_data(f"Лист1!J2:J{2+i}")
        abob = Reference(ws, min_col=1, min_row=2, max_row=2+i)
        chart.set_categories(abob)
        ws.add_chart(chart)
        wb.save(f"C:/prog_questionmark/test/data/second_zero/second_zero{message.chat.id}.xlsx")
        await bot.send_document(message.chat.id, open(f"C:/prog_questionmark/test/data/second_zero/second_zero{message.chat.id}.xlsx", 'rb'))
        #------------------------------------------
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["second_first"])
async def second_first(message : types.message):
    try:
        i2_1_input = message.text
        i2_1_cmd = []
        i2_1_cmd = i2_1_input.split("|")
        years_input_prep = i2_1_cmd[0]
        years_input = years_input_prep
        #years_input = input("Введи годы через пробел ")
        years_str = years_input.split(" ")
        years_str.remove("/second_first")
        #years_str = ['2004', '2005', '2006', '2007', '2008', '2009']
        years = []
        for value in years_str:
            years.append(float(value))
        #first_row_input = input("Верхний ряд через пробел ")
        first_row_str_prep = i2_1_cmd[1]
        first_row_str_prep = first_row_str_prep.replace(",", ".")
        first_row_str = first_row_str_prep.split(" ")
        #first_row_str = ['880', '890', '900', '1000']
        first_row = []
        second_row = []
        for value in first_row_str:
            first_row.append(float(value))
        #second_row_input= input("Нижний ряд через пробел ")
        second_row_str_prep = i2_1_cmd[2]
        second_row_str_prep = second_row_str_prep.replace(",", ".")
        second_row_str = second_row_str_prep.split(" ")
        #second_row_str = ['1020', '1040', '1080']
        for value in second_row_str:
            second_row.append(float(value))
        k_trans = second_row[0] / first_row[-1]
        #print(k_trans)
        row = []
        for i, value in enumerate(first_row):
            while i < len(first_row) - 1:
                row.append(round(value * k_trans, 2))
                break
        for value in second_row:
            row.append(value)
        await bot.send_message(message.from_user.id, f"Kперевода {k_trans}")
        await bot.send_message(message.from_user.id, "Второй способ")
        await bot.send_message(message.from_user.id, "Перый ряд")
        for i, value in enumerate(first_row):
            await bot.send_message(message.from_user.id, f"{years[i]} - {first_row[i] * 100 / first_row[-1]}")
        await bot.send_message(message.from_user.id, "Второй ряд")
        for i, value in enumerate(second_row):
            await bot.send_message(message.from_user.id, f"{years[i+len(first_row) - 1]} - {second_row[i] * 100 / second_row[0]}")
        await bot.send_message(message.from_user.id, f"Значения {row}")
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
    except:
        await bot.send_message(message.from_user.id, "Ошибка")


@dp.message_handler(commands=["second_second"])
async def second_second(message : types.message):
    try:
        i2_3_input = message.text
        i2_3_cmd = []
        i2_3_cmd = i2_3_input.split("|")
        years_input = i2_3_cmd[0]
        #years_input = input("Введи годы через пробел ")
        years_str = years_input.split(" ")
        print(years_str)
        years_str.remove("/second_second")
        print(years_str)
        years = []
        for value in years_str:
            years.append(float(value))
        first_row_input = i2_3_cmd[1]
        #first_row_input = input("Верхний ряд через пробел ")
        first_row_input_1 = first_row_input.replace(",", ".")
        first_row_str = first_row_input_1.split(" ")
        #first_row_str = ['101', '110', '125,5', '130,3', '140']
        first_row = []
        second_row = []
        for value in first_row_str:
            first_row.append(float(value))
        second_row_input = i2_3_cmd[2]
        #second_row_input = input("Нижний ряд через пробел ")
        second_row_input_1 = second_row_input.replace(",", ".")
        second_row_str = second_row_input_1.split(" ")
        #second_row_str = ['80', '85,2', '86,7', '89,3', '92,3']
        for value in second_row_str:
            second_row.append(float(value))
        perc_1 = []
        perc_2 = []
        for value in first_row:
            perc_1.append(round(value * 100 / first_row[0], 2))
        #print(f"Первый ряд - {perc_1}")
        for value in second_row:
            perc_2.append(round(value * 100 / second_row[0], 2))
        #print(f"Второй ряд - {perc_2}")
        t_max = 0
        t_min = 0
        perc_1_max = max(perc_1)
        perc_2_max = max(perc_2)
        if perc_1_max >= perc_2_max:
            t_max = perc_1_max / 100
            t_min = perc_2_max / 100
        else:
            t_max = perc_2_max / 100
            t_min = perc_1_max / 100
        k = t_max ** 0.25 / t_min ** 0.25
        await bot.send_message(message.from_user.id, "Абсолютные уровни рядов динамики")
        for i, value in enumerate(perc_1):
            await bot.send_message(message.from_user.id, f"{years[i]} ||| {perc_1[i]} ||| {perc_2[i]}")
        #print(f"Коэффициент опережения - {round(k, 2)}")
        await bot.send_message(message.from_user.id, f"Степень корня {len(perc_1) - 1}")
        await bot.send_message(message.from_user.id, f"Числитель {t_max}")
        await bot.send_message(message.from_user.id, f"Знаменатель {t_min}")
        #await bot.send_message(message.from_user.id, f"Коэффициент опережения  {round(k, 2)}")
        for i in range(len(first_row)):
            if perc_1[i] > perc_2[i]:
                await bot.send_message(message.from_user.id, f"Коп {years[i]} = {perc_1[i] / perc_2[i]}")
            else:
                await bot.send_message(message.from_user.id, f"Коп {years[i]} = {perc_2[i] / perc_1[i]}")
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["third"])
async def third(message : types.message):
    try:
        wb = load_workbook("C:/prog_questionmark/test/data/samples/list.xlsx")
        ws = wb.active 
        i_3_input = message.text
        i_3_cmd = []
        i_3_cmd = i_3_input.split("|")
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        months_1 = ["Январе", "Феврале", "Марте", "Апреле", "Мае", "Июне", "Июле", "Августе", "Сентябре", "Октябре", "Ноябре", "Декабре"]
        years_input_prep = i_3_cmd[0]
        years_input = years_input_prep.split(" ")
        years_input.remove("/third")
        #years_input = input("Введи годы через пробел ")
        #years_str = years_input.split(" ")
        years = []
        for value in years_input:
            years.append(float(value))
        first_row_input = i_3_cmd[1]
        #first_row_input = input("Верхний ряд через пробел ")
        first_row_str = first_row_input.split(" ")
        #first_row_str = ['322', '213', '266', '148', '286', '334', '338', '120', '146', '256', '146', '343']
        first_row = []
        second_row = []
        third_row = []
        for value in first_row_str:
            first_row.append(float(value))
        second_row_input = i_3_cmd[2]
        #second_row_input = input("Средний ряд через пробел ")
        second_row_str = second_row_input.split(" ")
        #second_row_str = ['100', '326', '233', '101', '282', '227', '169', '242', '349', '101', '225', '297']
        for value in second_row_str:
            second_row.append(float(value))
        third_row_input = i_3_cmd[3]
        #third_row_input= input("Нижний ряд через пробел ")
        third_row_str = third_row_input.split(" ")
        #third_row_str = ['141', '286', '102', '209', '185', '311', '254', '203', '149', '251', '168', '151']
        for value in third_row_str:
            third_row.append(float(value))
        sum_1 = sum(first_row)
        sum_2 = sum(second_row)
        sum_3 = sum(third_row)
        avg_1 = round(sum_1 / 12, 2)
        avg_2 = round(sum_2 / 12, 2)
        avg_3 = round(sum_3 / 12, 2)
        avg_row = []
        for i, value in enumerate(first_row):
            avg_row.append(round((value + second_row[i] + third_row[i]) / 3, 2))
        sum_4 = round(sum(avg_row), 2)
        avg_4 = round(sum(avg_row) / 12, 2)
        sezon = []
        await bot.send_message(message.from_user.id, f"Сумма первого года {sum_1}")
        await bot.send_message(message.from_user.id, f"Сумма второго года {sum_2}")
        await bot.send_message(message.from_user.id, f"Сумма третьего года {sum_3}")
        await bot.send_message(message.from_user.id, f"Сумма за все года {sum_4}")

        for i, value in enumerate(avg_row):
            sezon.append(round(value * 100 / avg_4, 2))
        for i, value in enumerate(months):
            await bot.send_message(message.from_user.id, f"{value} ||| {first_row[i]} ||| {second_row[i]} ||| {third_row[i]} ||| {avg_row[i]} ||| {sezon[i]}")
            ws[f"A{2+i}"].value = value
            ws[f"B{2+i}"].value = first_row[i]
            ws[f"C{2+i}"].value = second_row[i]
            ws[f"D{2+i}"].value = third_row[i]
            ws[f"E{2+i}"].value = avg_row[i]
            ws[f"F{2+i}"].value = sezon[i]
        await bot.send_message(message.from_user.id, f"Ср.уров.{avg_1} ||| {avg_2} ||| {avg_3} ||| {avg_4} ||| {round(sum(sezon) / 12, 0)}")
        for i, value in enumerate(sezon):
            if value < 100:
                await bot.send_message(message.from_user.id, f"В {months_1[i]} - уменьшилось на {round((sezon[i] - 100) * -1, 2)}%")
            else:
                await bot.send_message(message.from_user.id, f"В {months_1[i]} - увеличилось на {round(sezon[i] - 100, 2)}%")
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
        wb.save(f"C:/prog_questionmark/test/data/third/third{message.chat.id}.xlsx")
        await bot.send_document(message.chat.id, open(f"C:/prog_questionmark/test/data/third/third{message.chat.id}.xlsx", 'rb'))
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["fourth"])
async def fourth(message : types.message):
    try:
        wb = load_workbook("C:/prog_questionmark/test/data/samples/list3.xlsx")
        ws = wb.active
        i4_input = message.text
        i4_cmd = []
        i4_cmd = i4_input.split("|")
        first_row_input_prep = i4_cmd[0]
        #first_row_input = input("Верхний ряд через пробел ")
        first_row_str = first_row_input_prep.split(" ")
        first_row_str.remove("/fourth")
        #first_row_str = ['69,9', '69,8', '69,7', '68,5', '68,4', '67', '67', '66', '65', '64,1']
        first_row = []
        second_row = []
        third_row = []
        for value in first_row_str:
            value = value.replace(",", ".")
            first_row.append(float(value))
        second_row_input = i4_cmd[1]
        #second_row_input= input("Нижний ряд через пробел ")
        second_row_str = second_row_input.split(" ")
        #second_row_str = ['6', '4,8', '5,2', '3,56', '5,17', '4,36', '3,6', '3,21', '2,11', '1,98']
        count = 0
        for value in second_row_str:
            value = value.replace(",", ".")
            second_row.append(float(value))
            count += 1
        x_sqr = []
        for value in first_row:
            x_sqr.append(round(value * value, 4))
        y_sqr = []
        for value in second_row:
            y_sqr.append(round(value * value, 4))
        x_y = []
        for i, value in enumerate(first_row):
            x_y.append(round(value * second_row[i], 4))
        x_s_palkoi = round(sum(first_row) / count, 4)
        y_s_palkoi = round(sum(second_row) / count, 4)
        x_y_s_palkoi = round(sum(x_y) / count, 4)
        sum_y = sum(second_row)
        sum_x = sum(first_row)
        sum_x_sqr = sum(x_sqr)
        sum_y_sqr = sum(y_sqr)
        sum_x_y = sum(x_y)
        #sum_y = count * a + sum_x * b
        #sum_x_y = sum_x * a + sum_x_sqr * b
        A = np.array([[count, sum_x], [sum_x, sum_x_sqr]])
        B = np.array([[sum_y], [sum_x_y]])
        X = np.dot(np.linalg.inv(A), B)
        a = X[0]
        b = X[1]
        a = float(a)
        b = float(b)
        x_minus_avgx_sqr = []
        for i, value in enumerate(first_row):
            x_minus_avgx_sqr.append((value - x_s_palkoi) ** 2)
        ro_x = round((sum(x_minus_avgx_sqr) / count) ** 0.5, 4)
        y_minus_avgy_sqr = []
        for i, value in enumerate(second_row):
            y_minus_avgy_sqr.append((value - y_s_palkoi) ** 2)
        ro_y = round((sum(y_minus_avgy_sqr) / count) ** 0.5, 4)
        r = round((x_y_s_palkoi - x_s_palkoi * y_s_palkoi) / (ro_x * ro_y), 4)
        d = round(r ** 2, 2)
        func = []
        for i, value in enumerate(first_row):
            func.append(a + b*value)
        for i, value in enumerate(first_row):
            await bot.send_message(message.from_user.id, f"{i+1} ||| {x_sqr[i]} ||| {y_sqr[i]} ||| {x_y[i]} ||| {func[i]}")
        await bot.send_message(message.from_user.id, f"Х среднее {x_s_palkoi}")
        await bot.send_message(message.from_user.id, f"Y среднее {y_s_palkoi}")
        await bot.send_message(message.from_user.id, f"XY среднее {x_y_s_palkoi}")
        await bot.send_message(message.from_user.id, f"Сумма у {sum_y}")
        await bot.send_message(message.from_user.id, f"Сумма х {sum_x}")
        await bot.send_message(message.from_user.id, f"Сумма ху {sum_x_y}")
        await bot.send_message(message.from_user.id, f"Сумма х в квадрате {sum_x_sqr}")
        await bot.send_message(message.from_user.id, f"Сумма y в квадрате {sum_y_sqr}")
        await bot.send_message(message.from_user.id, f"Сумма Yx с палкой {sum(func)}")
        await bot.send_message(message.from_user.id, f"a {a}")
        await bot.send_message(message.from_user.id, f"b {b}")
        await bot.send_message(message.from_user.id, f"Числитель {sum(x_minus_avgx_sqr)}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count}")
        await bot.send_message(message.from_user.id, f"σ_x {ro_x}")
        await bot.send_message(message.from_user.id, f"Числитель {sum(y_minus_avgy_sqr)}")
        await bot.send_message(message.from_user.id, f"Знаменатель {count}")
        await bot.send_message(message.from_user.id, f"σ_y {ro_y}")
        await bot.send_message(message.from_user.id, f"Коэффициент корреляции r {r}")
        await bot.send_message(message.from_user.id, f"Коэффициент детерминации d {d}%")
        #b = (sum_x_y - ((sum_y - b * sum_x)) / count / sum_x_sqr)
        #a = (sum_y - b * sum_x) / count
        for i in range(count):
            ws[f"A{i+1}"].value = first_row[i]
            ws[f"B{i+1}"].value = second_row[i]
            ws[f"C{i+1}"].value = func[i]
            ws[f"E{i+1}"].value = 1+i
            ws[f"F{i+1}"].value = first_row[i]
            ws[f"G{i+1}"].value = second_row[i]
            ws[f"H{i+1}"].value = x_sqr[i]
            ws[f"I{i+1}"].value = y_sqr[i]
            ws[f"J{i+1}"].value = x_y[i]
            ws[f"K{i+1}"].value = func[i]

        chart = LineChart()
        chart.add_data(f"Лист1!B1:B{count}")
        chart.add_data(f"Лист1!C1:C{count}")
        abob = Reference(ws, min_col=1, min_row=1, max_row=count)
        chart.set_categories(abob)
        ws.add_chart(chart)
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
        wb.save(f"C:/prog_questionmark/test/data/fourth/fourth{message.chat.id}.xlsx")
        await bot.send_document(message.chat.id, open(f"C:/prog_questionmark/test/data/fourth/fourth{message.chat.id}.xlsx", 'rb'))
        await bot.send_message(message.from_user.id, "Графики нужно дооформить")
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["fifth"])
async def fifth(message : types.message):
    try:
        #print(message.chat.first_name)
        #print(f"{message.chat.first_name} {message.chat.last_name}")
        #i5_input = input("input")
        i5_input = message.text
        i5_input = i5_input.replace(",", ".")
        i5_str = i5_input.split(" ")
        i5_str.remove("/fifth")
        i5 = []
        for value in i5_str:
            i5.append(float(value))
        s = i5[0]
        n = i5[1]
        m = i5[2] * -1
        p = i5[3]
        v = i5[4] * -1
        s_end = s + n + m + p + v
        await bot.send_message(message.from_user.id, f"Численность населения на конец года {s_end}")
        s_avg = round((s_end + s) / 2, 2)
        await bot.send_message(message.from_user.id, f"Средняя численность населения {s_avg}")
        n_k = round(n * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент рождаемости {n_k}%")
        m_k = round(m * -1 * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент смертности {m_k}%")
        g_k = n_k - m_k
        await bot.send_message(message.from_user.id, f"Коэффициент естественного прироста {g_k}")
        oborot_k = round((n + m * -1) * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент оборота населения {oborot_k}%")
        eco_k = round((n_k - m_k) / (n_k + m_k), 2)
        await bot.send_message(message.from_user.id, f"Коэффициент экономичности воспроизводства {eco_k}")
        life_k = round(n / m, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент жизненности (депопуляции) населения {life_k}")
        #-------------------------------------------------
        plus_k = round(p * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент прибытия {plus_k}%")
        minus_k = round(v * -1 * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент выбытия {minus_k}%")
        mig_k = round((p + v) * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент миграции {mig_k}%")
        mig_int_k = round((p - v) * 1000 / s_avg, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент интенсивности миграционного оборота {mig_int_k}%")
        #print(p)
        #print(v)
        mig_eff_k = round((p + v) / (p - v) * 100, 2)
        await bot.send_message(message.from_user.id, f"Коэффициент эффективности миграции {mig_eff_k}%")
        last_k = g_k + mig_k
        await bot.send_message(message.from_user.id, f"Коэффициент общего прироста населения {last_k}%")
        cursor.execute("INSERT INTO `first` (`user_id`, `first_name`, `last_name`, `text`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
    except:
        await bot.send_message(message.from_user.id, "Ошибка")

@dp.message_handler(commands=["complaint"])
async def complaint(message : types.message):
    try:
        cursor.execute("INSERT INTO `complaints` (`user_id`, `first_name`, `last_name`, `complaint`) VALUES (?, ?, ?, ?)", (message.chat.id, message.chat.first_name, message.chat.last_name, message.text))
        conn.commit()
        await bot.send_message(message.from_user.id, "Твоя жалоба была принята")
    except:
        await bot.send_message(message.from_user.id, "И тут неудача, попробуй позже")

    
executor.start_polling(dp, skip_updates=True)